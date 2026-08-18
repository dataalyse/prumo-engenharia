"""
PRUMO ENGENHARIA - gerador de base sintetica para dashboard.

Diferenca para uma base "bonitinha": aqui nada sobe reto.
Cada obra tem uma linha do tempo de EVENTOS (dores e resolucoes) que
distorce produtividade, custo, prazo, seguranca e suprimentos - e essa
distorcao aparece nos graficos com causa nomeada.

Saidas:
    prumo.db      -> SQLite (fonte do Streamlit)
    prumo.xlsx    -> mesma base, 1 aba por tabela
    /csv/*.csv

Uso:
    python gerar_dados_prumo.py
"""

import os
import sqlite3
from datetime import date

import numpy as np
import pandas as pd

SEED = 7
rng = np.random.default_rng(SEED)

INICIO_BASE = date(2025, 3, 1)
N_MESES = 18                      # mar/2025 .. ago/2026
HOJE = pd.Timestamp("2026-08-01")  # ultimo mes com dado realizado
OUT_DIR = os.path.dirname(os.path.abspath(__file__))


# ======================================================================
# DIMENSOES
# ======================================================================
def dim_obra() -> pd.DataFrame:
    obras = [
        (1, "Residencial Vista Verde", "Residencial Vertical", "Angra dos Reis", 8_400,
         21_500_000, date(2025, 3, 1), 22, "Em Andamento", "Ricardo Almeida"),
        (2, "Edificio Corporativo Atlantica", "Comercial", "Rio de Janeiro", 12_600,
         38_900_000, date(2025, 5, 1), 24, "Em Andamento", "Fernanda Costa"),
        (3, "Condominio Mata Atlantica", "Residencial Horizontal", "Paraty", 5_200,
         13_200_000, date(2025, 9, 1), 16, "Em Andamento", "Marcos Tavares"),
        (4, "Loteamento Portal da Serra", "Infraestrutura", "Angra dos Reis", 31_000,
         9_800_000, date(2025, 3, 1), 14, "Concluida", "Ricardo Almeida"),
    ]
    df = pd.DataFrame(obras, columns=[
        "id_obra", "nome_obra", "tipo_obra", "cidade", "area_construida_m2",
        "orcamento_previsto", "data_inicio", "prazo_meses", "status_obra", "gerente_obra",
    ])
    df["data_termino_prevista"] = [
        (pd.Timestamp(i) + pd.DateOffset(months=int(p))).date()
        for i, p in zip(df["data_inicio"], df["prazo_meses"])
    ]
    df["custo_m2_orcado"] = (df["orcamento_previsto"] / df["area_construida_m2"]).round(2)
    return df


def dim_etapa() -> pd.DataFrame:
    etapas = [
        (1, "Servicos Preliminares", "Canteiro", 0.04, 1),
        (2, "Fundacoes", "Estrutura", 0.11, 2),
        (3, "Estrutura", "Estrutura", 0.24, 3),
        (4, "Alvenaria e Vedacao", "Vedacao", 0.12, 4),
        (5, "Instalacoes Hidraulicas", "Instalacoes", 0.08, 5),
        (6, "Instalacoes Eletricas", "Instalacoes", 0.09, 6),
        (7, "Revestimentos", "Acabamento", 0.14, 7),
        (8, "Esquadrias", "Acabamento", 0.06, 8),
        (9, "Pintura", "Acabamento", 0.05, 9),
        (10, "Paisagismo e Areas Comuns", "Externo", 0.04, 10),
        (11, "Limpeza e Entrega", "Externo", 0.03, 11),
    ]
    return pd.DataFrame(etapas, columns=[
        "id_etapa", "nome_etapa", "grupo_etapa", "peso_orcamento", "ordem_execucao"])


def dim_calendario() -> pd.DataFrame:
    datas = pd.date_range(INICIO_BASE, periods=N_MESES, freq="MS")
    df = pd.DataFrame({"_d": datas})
    df["id_tempo"] = df["_d"].dt.strftime("%Y%m").astype(int)
    df["ano_mes"] = df["_d"].dt.strftime("%Y-%m")
    df["ano"] = df["_d"].dt.year
    df["mes"] = df["_d"].dt.month
    meses_pt = {1: "jan", 2: "fev", 3: "mar", 4: "abr", 5: "mai", 6: "jun",
                7: "jul", 8: "ago", 9: "set", 10: "out", 11: "nov", 12: "dez"}
    df["nome_mes"] = df["mes"].map(meses_pt) + "/" + df["ano"].astype(str).str[2:]
    df["trimestre"] = "T" + df["_d"].dt.quarter.astype(str)
    df["estacao_chuva"] = df["mes"].isin([12, 1, 2, 3]).astype(int)
    df["data_ref"] = df["_d"].dt.date
    return df.drop(columns="_d")[[
        "id_tempo", "data_ref", "ano_mes", "ano", "mes", "nome_mes", "trimestre", "estacao_chuva"]]


def dim_fornecedor() -> pd.DataFrame:
    f = [
        (1, "Aco Forte Distribuidora", "Aco", "Rio de Janeiro", "A"),
        (2, "Vergalhao Sul Siderurgia", "Aco", "Volta Redonda", "A"),
        (3, "Concreteira Baia da Ilha", "Concreto", "Angra dos Reis", "A"),
        (4, "Ceramica Sul Fluminense", "Revestimento", "Volta Redonda", "B"),
        (5, "Eletro Supri Ltda", "Eletrica", "Rio de Janeiro", "B"),
        (6, "Hidro Master Materiais", "Hidraulica", "Niteroi", "B"),
        (7, "Madeireira Costa Verde", "Madeira", "Paraty", "C"),
        (8, "Esquadrias Aluminio RJ", "Esquadria", "Rio de Janeiro", "A"),
        (9, "Tintas Litoral", "Pintura", "Angra dos Reis", "C"),
        (10, "Impermeabiliza Rio", "Impermeabilizacao", "Rio de Janeiro", "C"),
    ]
    return pd.DataFrame(f, columns=[
        "id_fornecedor", "nome_fornecedor", "categoria", "cidade", "curva_abc"])


# ======================================================================
# LINHA DO TEMPO DE EVENTOS  (o coracao da base)
# ======================================================================
# fator_produtividade: multiplica o avanco fisico do mes (1.0 = plano)
# fator_custo:         multiplica o custo realizado do mes
# tipo:                'dor' ou 'resolucao'
EVENTOS = [
    # ---------------- Obra 1 - Vista Verde ----------------
    dict(id_obra=1, ano_mes="2025-08", tipo="dor", categoria="Suprimentos",
         titulo="Ruptura de estoque de aco",
         descricao="Fornecedor principal atrasou 42 dias. Concretagem de 3 lajes suspensa; "
                   "compra emergencial no mercado spot a preco 24% acima do cotado.",
         acao="", fator_produtividade=0.58, fator_custo=1.19, meses_efeito=2),
    dict(id_obra=1, ano_mes="2025-10", tipo="resolucao", categoria="Suprimentos",
         titulo="Segundo fornecedor de aco homologado",
         descricao="Contrato de fornecimento dividido entre dois fornecedores classe A. "
                   "Lead time medio caiu de 34 para 16 dias.",
         acao="Dual sourcing + contrato de fornecimento programado",
         fator_produtividade=1.16, fator_custo=0.97, meses_efeito=3),
    dict(id_obra=1, ano_mes="2025-12", tipo="dor", categoria="Clima",
         titulo="Chuvas de verao paralisam servicos externos",
         descricao="187 mm acima da media historica. 21 dias improdutivos acumulados em "
                   "alvenaria externa, impermeabilizacao e movimentacao de grua.",
         acao="", fator_produtividade=0.52, fator_custo=1.04, meses_efeito=3),
    dict(id_obra=1, ano_mes="2026-03", tipo="resolucao", categoria="Planejamento",
         titulo="Mutirao de recuperacao de cronograma",
         descricao="Antecipacao de servicos internos durante a chuva e reforco de efetivo. "
                   "Recuperados 6,1 p.p. de avanco em 3 meses.",
         acao="Replanejamento Last Planner + 2o turno em servicos internos",
         fator_produtividade=1.34, fator_custo=1.07, meses_efeito=3),

    # ---------------- Obra 2 - Atlantica ----------------
    dict(id_obra=2, ano_mes="2025-09", tipo="dor", categoria="Projeto",
         titulo="Revisao de projeto estrutural apos compatibilizacao",
         descricao="Interferencia entre estrutura e shafts hidraulicos detectada tardiamente. "
                   "Demolicao parcial de 2 pavimentos e reexecucao.",
         acao="", fator_produtividade=0.61, fator_custo=1.28, meses_efeito=3),
    dict(id_obra=2, ano_mes="2025-12", tipo="dor", categoria="Seguranca",
         titulo="Acidente com afastamento na fachada",
         descricao="Queda de nivel em plataforma de fachada. Obra interditada 5 dias uteis; "
                   "38 dias perdidos e reforco de protecao coletiva.",
         acao="", fator_produtividade=0.73, fator_custo=1.06, meses_efeito=1),
    dict(id_obra=2, ano_mes="2026-02", tipo="resolucao", categoria="Seguranca",
         titulo="Programa de protecao coletiva revisado",
         descricao="Linha de vida continua, treinamento NR-35 para 100% do efetivo de fachada. "
                   "Sem acidentes com afastamento nos 6 meses seguintes.",
         acao="Revisao de APR + NR-35 para todo o efetivo de altura",
         fator_produtividade=1.05, fator_custo=1.01, meses_efeito=6),
    dict(id_obra=2, ano_mes="2026-04", tipo="resolucao", categoria="Contrato",
         titulo="Aditivo contratual aprovado pelo cliente",
         descricao="Escopo adicional de fachada ventilada reconhecido: R$ 3,1 mi incorporados "
                   "ao orcamento. Desvio aparente de custo cai 8 p.p.",
         acao="Pleito formalizado com memoria de calculo e medicao",
         fator_produtividade=1.08, fator_custo=0.93, meses_efeito=5),

    # ---------------- Obra 3 - Mata Atlantica ----------------
    dict(id_obra=3, ano_mes="2025-12", tipo="dor", categoria="Clima",
         titulo="Temporal em Paraty interrompe terraplenagem",
         descricao="Escorregamento de talude no acesso da obra. Movimentacao de terra e "
                   "fundacoes paradas; retrabalho de contencao nao previsto.",
         acao="", fator_produtividade=0.44, fator_custo=1.15, meses_efeito=3),
    dict(id_obra=3, ano_mes="2026-03", tipo="dor", categoria="Licenciamento",
         titulo="Embargo ambiental por supressao vegetal",
         descricao="Obra embargada 26 dias por divergencia na area de supressao autorizada. "
                   "Efetivo mantido em canteiro sem producao.",
         acao="", fator_produtividade=0.21, fator_custo=1.09, meses_efeito=1),
    dict(id_obra=3, ano_mes="2026-05", tipo="resolucao", categoria="Licenciamento",
         titulo="Licenca retificada e obra desembargada",
         descricao="Compensacao ambiental acordada com o orgao e retomada com efetivo reforcado.",
         acao="Regularizacao junto ao INEA + plano de compensacao",
         fator_produtividade=1.42, fator_custo=1.05, meses_efeito=3),
    dict(id_obra=3, ano_mes="2026-07", tipo="dor", categoria="Qualidade",
         titulo="Reprovacao em massa na impermeabilizacao",
         descricao="Teste de estanqueidade reprovou 14 de 18 boxes. Fornecedor de manta "
                   "substituido e servico refeito.",
         acao="", fator_produtividade=0.82, fator_custo=1.12, meses_efeito=2),

    # ---------------- Obra 4 - Portal da Serra ----------------
    dict(id_obra=4, ano_mes="2025-07", tipo="dor", categoria="Licenciamento",
         titulo="Atraso na aprovacao do projeto de drenagem",
         descricao="Prefeitura reteve o projeto de drenagem por 48 dias. Pavimentacao "
                   "represada aguardando rede de aguas pluviais.",
         acao="", fator_produtividade=0.49, fator_custo=1.03, meses_efeito=2),
    dict(id_obra=4, ano_mes="2025-09", tipo="resolucao", categoria="Planejamento",
         titulo="Reordenacao de frentes de servico",
         descricao="Frentes de terraplenagem e meio-fio antecipadas enquanto a drenagem "
                   "tramitava. Cronograma recuperado sem custo adicional relevante.",
         acao="Resequenciamento da EAP com frentes independentes",
         fator_produtividade=1.28, fator_custo=1.0, meses_efeito=3),
    dict(id_obra=4, ano_mes="2026-01", tipo="dor", categoria="Clima",
         titulo="Chuva na fase de pavimentacao asfaltica",
         descricao="Janelas secas insuficientes para imprimacao. 11 dias improdutivos.",
         acao="", fator_produtividade=0.66, fator_custo=1.05, meses_efeito=2),
]


def fato_eventos(cal: pd.DataFrame) -> pd.DataFrame:
    df = pd.DataFrame(EVENTOS)
    df.insert(0, "id_evento", range(1, len(df) + 1))
    df = df.merge(cal[["id_tempo", "ano_mes", "data_ref"]], on="ano_mes", how="left")
    return df[["id_evento", "id_obra", "id_tempo", "data_ref", "ano_mes", "tipo",
               "categoria", "titulo", "descricao", "acao", "fator_produtividade",
               "fator_custo", "meses_efeito"]]


def _fatores_mensais(id_obra: int, ano_mes_list: list) -> tuple:
    """Expande os eventos no tempo (com decaimento) e devolve as series de fatores."""
    n = len(ano_mes_list)
    idx = {am: i for i, am in enumerate(ano_mes_list)}
    f_prod = np.ones(n)
    f_custo = np.ones(n)
    for ev in EVENTOS:
        if ev["id_obra"] != id_obra or ev["ano_mes"] not in idx:
            continue
        i0 = idx[ev["ano_mes"]]
        dur = ev["meses_efeito"]
        for k in range(dur):
            i = i0 + k
            if i >= n:
                break
            # o impacto decai linearmente ate o fim da janela do evento
            peso = 1 - (k / dur) * 0.55
            f_prod[i] *= 1 + (ev["fator_produtividade"] - 1) * peso
            f_custo[i] *= 1 + (ev["fator_custo"] - 1) * peso
    return f_prod, f_custo


# ======================================================================
# CURVA S + AVANCO FISICO
# ======================================================================
def _curva_s(n: int, alpha=2.4, beta=2.2) -> np.ndarray:
    x = np.linspace(0, 1, n + 1)[1:]
    def cdf(t):
        u = np.linspace(1e-6, max(t, 1e-6), 300)
        return np.trapezoid(u ** (alpha - 1) * (1 - u) ** (beta - 1), u)
    b = np.array([cdf(t) for t in x])
    return b / b[-1]


def _ruido_ar1(n: int, phi=0.55, sigma=0.11) -> np.ndarray:
    """Ruido autocorrelacionado: um mes ruim tende a puxar o seguinte.
    Ruido branco produziria serrilhado irreal; AR(1) produz ondas."""
    e = rng.normal(0, sigma, n)
    x = np.zeros(n)
    for i in range(1, n):
        x[i] = phi * x[i - 1] + e[i]
    return x


def fato_avanco(obras, cal) -> pd.DataFrame:
    linhas = []
    for _, o in obras.iterrows():
        ini = pd.Timestamp(o["data_inicio"])
        fim = ini + pd.DateOffset(months=int(o["prazo_meses"]) - 1)
        jan = cal[(pd.to_datetime(cal["data_ref"]) >= ini) &
                  (pd.to_datetime(cal["data_ref"]) <= fim)].reset_index(drop=True)
        n = len(jan)
        if n == 0:
            continue

        prev_acum = _curva_s(int(o["prazo_meses"]))[:n]
        prev_mes = np.diff(prev_acum, prepend=0)

        f_prod, _ = _fatores_mensais(int(o["id_obra"]), jan["ano_mes"].tolist())
        # chuva de verao penaliza mesmo sem evento nomeado
        f_chuva = np.where(jan["estacao_chuva"].values == 1, 0.88, 1.0)
        ruido = 1 + _ruido_ar1(n)
        prod = np.clip(f_prod * f_chuva * ruido, 0.10, 1.75)

        real_mes = prev_mes * prod
        real_acum = np.clip(np.cumsum(real_mes), 0, 1.0)

        n_real = int((pd.to_datetime(jan["data_ref"]) <= HOJE).sum())
        if o["status_obra"] == "Concluida":
            real_acum = real_acum / real_acum[-1]
            real_mes = np.diff(real_acum, prepend=0)
            n_real = n

        for i, r in jan.iterrows():
            tem = i < n_real
            linhas.append({
                "id_obra": int(o["id_obra"]),
                "id_tempo": int(r["id_tempo"]),
                "data_ref": r["data_ref"],
                "ano_mes": r["ano_mes"],
                "avanco_previsto_mes": round(float(prev_mes[i]), 4),
                "avanco_real_mes": round(float(real_mes[i]), 4) if tem else None,
                "avanco_previsto_acum": round(float(prev_acum[i]), 4),
                "avanco_real_acum": round(float(real_acum[i]), 4) if tem else None,
                "indice_produtividade": round(float(prod[i]), 3) if tem else None,
            })
    df = pd.DataFrame(linhas)
    df["desvio_prazo_pp"] = ((df["avanco_real_acum"] - df["avanco_previsto_acum"]) * 100).round(2)
    return df


# ======================================================================
# CUSTO
# ======================================================================
def fato_custo(obras, etapas, avanco) -> pd.DataFrame:
    naturezas = ["Material", "Mao de Obra", "Equipamento", "Servico Terceirizado", "Indireto"]
    peso_nat = np.array([0.45, 0.30, 0.08, 0.13, 0.04])
    linhas = []

    for _, o in obras.iterrows():
        av = avanco[avanco["id_obra"] == o["id_obra"]].reset_index(drop=True)
        n = len(av)
        if n == 0:
            continue
        _, f_custo = _fatores_mensais(int(o["id_obra"]), av["ano_mes"].tolist())
        deriva = 1 + np.linspace(0, 0.05, n)  # inflacao de insumos ao longo da obra

        for _, et in etapas.iterrows():
            centro = (et["ordem_execucao"] - 0.5) / len(etapas)
            pos = np.linspace(0, 1, n)
            w = np.exp(-0.5 * ((pos - centro) / 0.22) ** 2)
            if w.sum() == 0:
                continue
            w = w / w.sum()
            orc_etapa = o["orcamento_previsto"] * et["peso_orcamento"]
            mult_etapa = 1.05 if et["grupo_etapa"] == "Estrutura" else 1.0

            for i, r in av.iterrows():
                orc_mes = orc_etapa * w[i]
                if orc_mes < 500:
                    continue
                tem = pd.notna(r["avanco_real_acum"])
                for nat, pn in zip(naturezas, peso_nat):
                    orcado = orc_mes * pn
                    realizado = None
                    if tem:
                        fator = f_custo[i] * mult_etapa * deriva[i] * rng.normal(1.0, 0.06)
                        if nat == "Material":
                            fator *= rng.normal(1.02, 0.04)
                        if nat == "Mao de Obra" and r["indice_produtividade"] and r["indice_produtividade"] < 0.8:
                            # efetivo parado em canteiro custa igual e produz menos
                            fator *= 1.12
                        realizado = orcado * max(fator, 0.55)
                    linhas.append({
                        "id_obra": int(o["id_obra"]),
                        "id_etapa": int(et["id_etapa"]),
                        "id_tempo": int(r["id_tempo"]),
                        "data_ref": r["data_ref"],
                        "ano_mes": r["ano_mes"],
                        "natureza_custo": nat,
                        "custo_orcado": round(float(orcado), 2),
                        "custo_realizado": None if realizado is None else round(float(realizado), 2),
                    })
    df = pd.DataFrame(linhas)
    df["desvio_valor"] = (df["custo_realizado"] - df["custo_orcado"]).round(2)
    df["desvio_pct"] = (df["desvio_valor"] / df["custo_orcado"] * 100).round(2)
    return df


# ======================================================================
# SUPRIMENTOS
# ======================================================================
def fato_compras(obras, forns, cal) -> pd.DataFrame:
    """OTIF piora na ruptura de aco (obra 1, ago-set/25) e melhora apos o dual sourcing."""
    linhas = []
    pid = 0
    for _, o in obras.iterrows():
        ini = pd.Timestamp(o["data_inicio"])
        fim = min(ini + pd.DateOffset(months=int(o["prazo_meses"]) - 1), HOJE)
        jan = cal[(pd.to_datetime(cal["data_ref"]) >= ini) &
                  (pd.to_datetime(cal["data_ref"]) <= fim)]
        for _, r in jan.iterrows():
            for _ in range(int(rng.integers(4, 10))):
                pid += 1
                f = forns.sample(1, random_state=int(rng.integers(0, 1e6))).iloc[0]
                cotado = float(rng.lognormal(10.6, 0.72))

                saving = float(np.clip(rng.normal(0.062, 0.042), -0.04, 0.22))
                base = {"A": 0.80, "B": 0.94, "C": 1.14}[f["curva_abc"]]

                # crise do aco na obra 1
                crise = (o["id_obra"] == 1 and f["categoria"] == "Aco"
                         and r["ano_mes"] in ("2025-08", "2025-09"))
                if crise:
                    base *= 2.3
                    saving = -0.24  # compra spot: pagou acima do cotado
                # apos o dual sourcing o aco normaliza
                if (o["id_obra"] == 1 and f["categoria"] == "Aco"
                        and r["ano_mes"] >= "2025-10"):
                    base *= 0.72
                    saving = float(np.clip(rng.normal(0.11, 0.03), 0, 0.25))
                # manta reprovada na obra 3
                if (o["id_obra"] == 3 and f["categoria"] == "Impermeabilizacao"
                        and r["ano_mes"] >= "2026-07"):
                    base *= 1.5

                lp = int(rng.integers(5, 26))
                lr = max(1, int(round(lp * rng.normal(base, 0.20))))
                linhas.append({
                    "id_pedido": pid, "id_obra": int(o["id_obra"]),
                    "id_fornecedor": int(f["id_fornecedor"]),
                    "id_tempo": int(r["id_tempo"]), "data_pedido": r["data_ref"],
                    "ano_mes": r["ano_mes"], "categoria": f["categoria"],
                    "valor_cotado": round(cotado, 2),
                    "valor_comprado": round(cotado * (1 - saving), 2),
                    "saving_valor": round(cotado * saving, 2),
                    "saving_pct": round(saving * 100, 2),
                    "lead_time_previsto_dias": lp, "lead_time_real_dias": lr,
                    "entrega_otif": int(lr <= lp),
                })
    return pd.DataFrame(linhas)


# ======================================================================
# MAO DE OBRA / SEGURANCA / QUALIDADE / VENDAS
# ======================================================================
def fato_mao_obra(obras, avanco) -> pd.DataFrame:
    linhas = []
    for _, o in obras.iterrows():
        av = avanco[(avanco["id_obra"] == o["id_obra"]) & avanco["avanco_real_acum"].notna()]
        for _, r in av.iterrows():
            base = o["area_construida_m2"] / 180
            ritmo = max(r["avanco_previsto_mes"], 0.005)   # efetivo segue o PLANO
            ip = r["indice_produtividade"] or 1.0
            proprio = int(max(6, rng.normal(base * ritmo * 9, 3)))
            terc = int(max(2, proprio * rng.normal(0.55, 0.12)))
            total = proprio + terc
            hh = total * 176
            # obra atrasada compensa com hora extra
            he_pct = np.clip(rng.normal(0.05 + max(0, (1 - ip)) * 0.16, 0.02), 0, 0.32)
            linhas.append({
                "id_obra": int(o["id_obra"]), "id_tempo": int(r["id_tempo"]),
                "data_ref": r["data_ref"], "ano_mes": r["ano_mes"],
                "efetivo_proprio": proprio, "efetivo_terceirizado": terc,
                "efetivo_total": total,
                "hh_normais": round(hh, 1), "hh_extras": round(hh * float(he_pct), 1),
                "custo_hh_medio": round(float(rng.normal(28.5, 2.0)), 2),
                "absenteismo_pct": round(float(np.clip(rng.normal(4.1, 1.5), 0.5, 12)), 2),
                "turnover_pct": round(float(np.clip(rng.normal(5.8, 2.2), 0.5, 18)), 2),
                "indice_produtividade": r["indice_produtividade"],
            })
    return pd.DataFrame(linhas)


def fato_seguranca(mo) -> pd.DataFrame:
    linhas = []
    for _, r in mo.iterrows():
        hht = r["hh_normais"] + r["hh_extras"]
        lam = hht / 1e6 * 11
        # acidente da fachada: obra 2, dez/25
        acidente = (r["id_obra"] == 2 and r["ano_mes"] == "2025-12")
        # programa de protecao coletiva zera afastamentos na obra 2 a partir de fev/26
        protegido = (r["id_obra"] == 2 and r["ano_mes"] >= "2026-02")
        ca = 2 if acidente else (0 if protegido else int(rng.poisson(lam * 0.35)))
        sa = int(rng.poisson(lam * (0.4 if protegido else 0.65)))
        dias = 38 if acidente else (int(ca * rng.integers(3, 26)) if ca else 0)
        linhas.append({
            "id_obra": int(r["id_obra"]), "id_tempo": int(r["id_tempo"]),
            "data_ref": r["data_ref"], "ano_mes": r["ano_mes"],
            "hht": round(float(hht), 1),
            "acidentes_com_afastamento": ca, "acidentes_sem_afastamento": sa,
            "dias_perdidos": dias,
            "dds_realizados": int(rng.integers(16, 23)),
            "inspecoes_realizadas": int(rng.integers(2, 7)),
            "nao_conformidades_epi": int(rng.poisson(3.4 if acidente else 2.2)),
        })
    return pd.DataFrame(linhas)


def fato_qualidade(obras, etapas, avanco) -> pd.DataFrame:
    linhas = []
    for _, o in obras.iterrows():
        av = avanco[(avanco["id_obra"] == o["id_obra"]) & avanco["avanco_real_acum"].notna()]
        for _, r in av.iterrows():
            for _, et in etapas.sample(3, random_state=int(rng.integers(0, 1e6))).iterrows():
                insp = int(rng.integers(4, 19))
                taxa = float(np.clip(rng.normal(0.84, 0.08), 0.4, 0.99))
                # reprovacao em massa da impermeabilizacao (obra 3, jul-ago/26)
                impermeab = (o["id_obra"] == 3 and r["ano_mes"] >= "2026-07"
                             and et["grupo_etapa"] in ("Instalacoes", "Acabamento"))
                if impermeab:
                    taxa = 0.22
                    insp = 18
                # retrabalho estrutural da obra 2
                if o["id_obra"] == 2 and r["ano_mes"] in ("2025-09", "2025-10", "2025-11") \
                        and et["grupo_etapa"] == "Estrutura":
                    taxa = 0.41
                ok = int(round(insp * taxa))
                linhas.append({
                    "id_obra": int(o["id_obra"]), "id_etapa": int(et["id_etapa"]),
                    "id_tempo": int(r["id_tempo"]), "data_ref": r["data_ref"],
                    "ano_mes": r["ano_mes"],
                    "inspecoes_fvs": insp, "aprovados_primeira": ok,
                    "reprovados": insp - ok,
                    "custo_retrabalho": round(float((insp - ok) * rng.lognormal(7.5, 0.55)), 2),
                    "perda_material_pct": round(float(np.clip(rng.normal(6.5, 2.6), 0.5, 18)), 2),
                })
    return pd.DataFrame(linhas)


def fato_vendas(obras, cal) -> pd.DataFrame:
    unid = {1: 96, 2: 48, 3: 32, 4: 180}
    pm2 = {1: 8200, 2: 11500, 3: 9400, 4: 1250}
    linhas = []
    for _, o in obras.iterrows():
        idb = int(o["id_obra"])
        ini = pd.Timestamp(o["data_inicio"])
        fim = min(ini + pd.DateOffset(months=int(o["prazo_meses"]) - 1), HOJE)
        jan = cal[(pd.to_datetime(cal["data_ref"]) >= ini) &
                  (pd.to_datetime(cal["data_ref"]) <= fim)]
        estoque, acum = unid[idb], 0
        ticket = o["area_construida_m2"] / unid[idb] * pm2[idb]
        for _, r in jan.iterrows():
            if estoque <= 0:
                break
            vso = rng.normal(0.075, 0.028)
            # noticia de acidente e embargo respingam na venda
            if (idb == 2 and r["ano_mes"] in ("2025-12", "2026-01")) or \
               (idb == 3 and r["ano_mes"] in ("2026-03", "2026-04")):
                vso *= 0.45
            v = min(int(max(0, round(estoque * vso))), estoque)
            dist = int(rng.binomial(max(acum, 0), 0.008))
            estoque = estoque - v + dist
            acum += v - dist
            linhas.append({
                "id_obra": idb, "id_tempo": int(r["id_tempo"]), "data_ref": r["data_ref"],
                "ano_mes": r["ano_mes"], "unidades_totais": unid[idb],
                "unidades_vendidas_mes": v, "unidades_vendidas_acum": acum,
                "estoque_unidades": estoque, "distratos_mes": dist,
                "ticket_medio": round(float(ticket), 2),
                "vgv_vendido_mes": round(float(v * ticket), 2),
                "inadimplencia_pct": round(float(np.clip(rng.normal(3.8, 1.4), 0.2, 12)), 2),
            })
    df = pd.DataFrame(linhas)
    df["vso_pct"] = (df["unidades_vendidas_mes"] /
                     (df["estoque_unidades"] + df["unidades_vendidas_mes"]) * 100).round(2)
    return df


# ======================================================================
def dicionario() -> pd.DataFrame:
    l = [
        ("dim_obra", "Dimensao", "Cadastro das 4 obras. Chave: id_obra"),
        ("dim_etapa", "Dimensao", "EAP. peso_orcamento soma 1,00"),
        ("dim_calendario", "Dimensao", "mar/2025 a ago/2026. estacao_chuva=1 em dez-mar"),
        ("dim_fornecedor", "Dimensao", "10 fornecedores com curva ABC"),
        ("fato_eventos", "Fato", "Linha do tempo de dores e resolucoes por obra - dirige toda a base"),
        ("fato_avanco_fisico", "Fato", "Curva S com produtividade real; indice_produtividade 1,0 = plano"),
        ("fato_custo", "Fato", "Orcado x realizado por obra/etapa/mes/natureza"),
        ("fato_compras", "Fato", "Pedidos: saving, lead time, OTIF"),
        ("fato_mao_obra", "Fato", "Efetivo, HH, hora extra, absenteismo, turnover"),
        ("fato_seguranca", "Fato", "Acidentes, HHT, dias perdidos, DDS, inspecoes"),
        ("fato_qualidade", "Fato", "FVS, retrabalho, perda de material"),
        ("fato_vendas", "Fato", "VSO, VGV, estoque, distratos, inadimplencia"),
    ]
    n = [
        ("PREMISSA", "-", "Base 100% sintetica (numpy, seed=7). Nao representa obra real."),
        ("PREMISSA", "-", "Avanco real = avanco previsto x indice de produtividade do mes."),
        ("PREMISSA", "-", "Produtividade = fator dos eventos x sazonalidade de chuva x ruido AR(1)."),
        ("PREMISSA", "-", "Ruido AR(1) (phi=0,55): mes ruim contamina o seguinte, como no canteiro."),
        ("PREMISSA", "-", "Efetivo segue o cronograma PREVISTO: por isso mes improdutivo encarece o HH."),
        ("PREMISSA", "-", "Meses futuros ficam com realizado NULO de proposito."),
        ("CALCULO", "-", "Taxa Frequencia = acidentes com afastamento / HHT x 1.000.000"),
        ("CALCULO", "-", "Taxa Gravidade = dias perdidos / HHT x 1.000.000"),
        ("CALCULO", "-", "Desvio de prazo (p.p.) = (real_acum - previsto_acum) x 100"),
        ("CALCULO", "-", "VSO = vendidas_mes / (estoque + vendidas_mes) x 100"),
    ]
    return pd.concat([pd.DataFrame(l, columns=["tabela", "tipo", "descricao"]),
                      pd.DataFrame(n, columns=["tabela", "tipo", "descricao"])], ignore_index=True)


def exportar(tabelas: dict):
    db = os.path.join(OUT_DIR, "prumo.db")
    if os.path.exists(db):
        os.remove(db)
    con = sqlite3.connect(db)
    for nome, df in tabelas.items():
        df.to_sql(nome, con, index=False, if_exists="replace")
    cur = con.cursor()
    for t, df in tabelas.items():
        if t.startswith("fato_"):
            if "id_obra" in df.columns:
                cur.execute(f"CREATE INDEX IF NOT EXISTS ix_{t}_obra ON {t}(id_obra)")
            if "id_tempo" in df.columns:
                cur.execute(f"CREATE INDEX IF NOT EXISTS ix_{t}_tempo ON {t}(id_tempo)")
    con.commit()
    con.close()

    cdir = os.path.join(OUT_DIR, "csv")
    os.makedirs(cdir, exist_ok=True)
    for nome, df in tabelas.items():
        df.to_csv(os.path.join(cdir, f"{nome}.csv"), index=False, encoding="utf-8-sig")

    xlsx = os.path.join(OUT_DIR, "prumo.xlsx")
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        for nome, df in tabelas.items():
            df.to_excel(xw, sheet_name=nome[:31], index=False)
    _formatar(xlsx)
    return db, xlsx, cdir


def _formatar(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1B2430")
    fcab = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fcor = Font(name="Arial", size=10)
    for ws in wb.worksheets:
        for c in ws[1]:
            c.fill, c.font = fill, fcab
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = fcor
                if isinstance(c.value, float):
                    c.number_format = "#,##0.00"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col[:150] if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 11), 40)
    wb.save(path)


def main():
    o, e, c, f = dim_obra(), dim_etapa(), dim_calendario(), dim_fornecedor()
    ev = fato_eventos(c)
    av = fato_avanco(o, c)
    cu = fato_custo(o, e, av)
    co = fato_compras(o, f, c)
    mo = fato_mao_obra(o, av)
    sg = fato_seguranca(mo)
    ql = fato_qualidade(o, e, av)
    vd = fato_vendas(o, c)

    tabelas = {
        "dim_obra": o, "dim_etapa": e, "dim_calendario": c, "dim_fornecedor": f,
        "fato_eventos": ev, "fato_avanco_fisico": av, "fato_custo": cu,
        "fato_compras": co, "fato_mao_obra": mo, "fato_seguranca": sg,
        "fato_qualidade": ql, "fato_vendas": vd, "dicionario_dados": dicionario(),
    }
    db, xlsx, cdir = exportar(tabelas)
    print("PRUMO ENGENHARIA - base gerada\n")
    for n, df in tabelas.items():
        print(f"  {n:<22} {len(df):>7,} linhas | {len(df.columns)} colunas")
    print(f"\nSQLite: {db}\nExcel : {xlsx}\nCSVs  : {cdir}")


if __name__ == "__main__":
    main()
